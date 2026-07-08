from __future__ import annotations

import re
import tempfile
import os
import base64
import html
import time
from pathlib import Path

import requests
import streamlit as st

from src.ai_client import GeminiError
from src.config import APP_NAME, DEFAULT_MODEL, DEFAULT_PROVIDER, GEMINI_API_KEY, OUTPUTS_DIR
from src.pipeline import run_pipeline
from src.usage_tracker import append_usage_row, build_usage_row

st.set_page_config(page_title=APP_NAME, layout="wide")

BASE_DIR = Path(__file__).parent
LOGO_PATH = BASE_DIR / "assets" / "pw_logo.png"
ALLOWED_EMAILS_PATH = BASE_DIR / "assets" / "allowed_emails.txt"
DEFAULT_ALLOWED_EMAILS_WORKBOOK = Path.home() / "Downloads" / "App Allowed Emails.xlsx"
DEFAULT_ALLOWED_EMAILS_SHEET_URL = "https://docs.google.com/spreadsheets/d/1ZpHOOYUVL_uz6MZ_yitrdUuSscLUNygCvRsh9-xhwa0/edit?usp=sharing"
EMAIL_RE = re.compile(r"[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}", re.IGNORECASE)
GOOGLE_SHEET_ID_RE = re.compile(r"/spreadsheets/d/([a-zA-Z0-9-_]+)")
SESSION_TIMEOUT_SECONDS = 7 * 24 * 60 * 60


def _configured_secret(name: str) -> str:
    try:
        return str(st.secrets.get(name, "")).strip()
    except Exception:
        return ""


def _configured_google_sheet_url() -> str:
    return (
        os.getenv("ALLOWED_EMAILS_GOOGLE_SHEET_URL", "").strip()
        or _configured_secret("allowed_emails_google_sheet_url")
        or DEFAULT_ALLOWED_EMAILS_SHEET_URL
    )


def _google_sheet_csv_url(sheet_url: str) -> str:
    match = GOOGLE_SHEET_ID_RE.search(sheet_url)
    if match:
        sheet_id = match.group(1)
    elif re.fullmatch(r"[a-zA-Z0-9-_]+", sheet_url):
        sheet_id = sheet_url
    else:
        return ""
    return f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv"


def _load_emails_from_google_sheet(sheet_url: str) -> set[str]:
    csv_url = _google_sheet_csv_url(sheet_url)
    if not csv_url:
        return set()

    response = requests.get(csv_url, timeout=10)
    response.raise_for_status()

    if "text/html" in response.headers.get("content-type", "").lower():
        return set()

    return {match.lower() for match in EMAIL_RE.findall(response.text)}


def _allowed_email_workbook_candidates() -> list[Path]:
    configured_path = os.getenv("ALLOWED_EMAILS_WORKBOOK", "").strip()
    try:
        configured_path = configured_path or _configured_secret("allowed_emails_workbook")
    except Exception:
        pass

    paths = [
        Path(configured_path).expanduser() if configured_path else None,
        BASE_DIR / "assets" / "App Allowed Emails.xlsx",
        DEFAULT_ALLOWED_EMAILS_WORKBOOK,
    ]

    unique_paths: list[Path] = []
    seen: set[str] = set()
    for path in paths:
        if not path:
            continue
        key = str(path).lower()
        if key not in seen:
            unique_paths.append(path)
            seen.add(key)
    return unique_paths


def _load_emails_from_workbook(path: Path) -> set[str]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    emails: set[str] = set()
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows(values_only=True):
            for value in row:
                if isinstance(value, str):
                    emails.update(match.lower() for match in EMAIL_RE.findall(value))
    return emails


def _load_allowed_emails() -> set[str]:
    sheet_url = _configured_google_sheet_url()
    if sheet_url:
        try:
            sheet_emails = _load_emails_from_google_sheet(sheet_url)
            if sheet_emails:
                return sheet_emails
        except requests.RequestException:
            pass

    emails: set[str] = set()
    for workbook_path in _allowed_email_workbook_candidates():
        if workbook_path.exists():
            emails.update(_load_emails_from_workbook(workbook_path))
            break

    if ALLOWED_EMAILS_PATH.exists():
        emails.update(
            match.lower()
            for match in EMAIL_RE.findall(ALLOWED_EMAILS_PATH.read_text(encoding="utf-8"))
        )
    return emails


def _google_auth_configured() -> bool:
    try:
        google_auth = st.secrets.get("auth", {}).get("google", {})
        client_id = str(google_auth.get("client_id", "")).strip()
        client_secret = str(google_auth.get("client_secret", "")).strip()
    except Exception:
        return False

    placeholder_values = ("", "your-google-oauth-client-id", "your-google-oauth-client-secret")
    return (
        client_id.endswith(".apps.googleusercontent.com")
        and client_secret not in placeholder_values
        and not client_id.startswith("your-google-oauth-client-id")
    )


def _current_user_email() -> str:
    email = ""
    try:
        email = st.user.get("email", "")
    except Exception:
        email = ""
    return str(email).strip().lower()


def _usd_to_inr_rate() -> float:
    configured = os.getenv("USD_TO_INR_RATE", "").strip() or _configured_secret("usd_to_inr_rate")
    try:
        return float(configured)
    except (TypeError, ValueError):
        return 83.0


def _google_access_token() -> str:
    try:
        tokens = st.user.get("tokens", {})
    except Exception:
        tokens = {}
    try:
        return str(tokens.get("access", "")).strip()
    except Exception:
        return str(getattr(tokens, "access", "") or "").strip()


def _track_usage_for_result(filename: str, result) -> str:
    metadata = result.metadata or {}
    row = build_usage_row(
        app_name=APP_NAME,
        email=_current_user_email(),
        filename=filename,
        input_unit=str(metadata.get("input_unit") or "slide/page"),
        count=int(metadata.get("active_slide_count") or 0),
        model=str(metadata.get("model") or DEFAULT_MODEL),
        metadata=metadata,
        image_model=IMAGE_MODEL,
        usd_to_inr=_usd_to_inr_rate(),
    )
    return append_usage_row(
        row,
        secrets=st.secrets,
        local_path=OUTPUTS_DIR / "usage_tracking.csv",
        user_access_token=_google_access_token(),
    )


def _logout_user() -> None:
    if bool(getattr(st.user, "is_logged_in", False)):
        st.logout()
    st.rerun()


def _auth_session_expired() -> bool:
    """Expire signed-in users 7 days after the identity token was issued."""
    try:
        issued_at = int(st.user.get("iat", 0) or 0)
    except (TypeError, ValueError):
        return False

    return issued_at > 0 and time.time() - issued_at >= SESSION_TIMEOUT_SECONDS


def _active_theme() -> str:
    theme = st.session_state.get("ui_theme", "Light")
    return "Dark" if theme == "Dark" else "Light"


def _render_theme_picker(key: str) -> None:
    current = _active_theme()
    st.radio(
        "Appearance",
        ["Light", "Dark"],
        index=0 if current == "Light" else 1,
        key=key,
        horizontal=True,
    )
    if st.session_state.get(key) != st.session_state.get("ui_theme"):
        st.session_state.ui_theme = st.session_state[key]
        st.rerun()


def _render_global_styles() -> None:
    is_dark = _active_theme() == "Dark"
    theme_vars = """
            --page: #f6f8fb;
            --surface: #ffffff;
            --surface-muted: #f1f5f9;
            --border: #d7dee8;
            --border-strong: #aebacc;
            --text: #111827;
            --text-soft: #4b5563;
            --text-muted: #6b7280;
            --brand: #123c69;
            --brand-strong: #0b2c4f;
            --brand-soft: #eaf3ff;
            --warning-bg: #fffbeb;
            --error-bg: #fef2f2;
            --success-bg: #ecfdf5;
            --shadow: rgba(15, 23, 42, 0.08);
    """
    if is_dark:
        theme_vars = """
            --page: #0f172a;
            --surface: #111827;
            --surface-muted: #1f2937;
            --border: #334155;
            --border-strong: #475569;
            --text: #f8fafc;
            --text-soft: #cbd5e1;
            --text-muted: #94a3b8;
            --brand: #60a5fa;
            --brand-strong: #93c5fd;
            --brand-soft: #172554;
            --warning-bg: #3b2f0b;
            --error-bg: #3f1218;
            --success-bg: #052e22;
            --shadow: rgba(0, 0, 0, 0.25);
        """

    st.markdown(
        """
        <style>
        :root {
__THEME_VARS__
        }
        .stApp {
            background: var(--page);
            color: var(--text);
            font-family: "Segoe UI", Arial, sans-serif;
        }
        [data-testid="stHeader"] {
            background: var(--surface);
            border-bottom: 1px solid var(--border);
        }
        [data-testid="stMainBlockContainer"] {
            max-width: 960px;
            padding-top: 1.5rem;
            padding-bottom: 3rem;
        }
        [data-testid="stSidebar"] {
            background: var(--surface);
            border-right: 1px solid var(--border);
        }
        h1, h2, h3, h4, h5, h6,
        p, label, span, small,
        [data-testid="stMarkdownContainer"],
        [data-testid="stMarkdownContainer"] * {
            color: var(--text);
        }
        [data-testid="stCaptionContainer"],
        [data-testid="stCaptionContainer"] * {
            color: var(--text-muted) !important;
        }
        .app-header {
            display: flex;
            align-items: center;
            justify-content: space-between;
            gap: 1rem;
            margin-bottom: 1rem;
            padding: 1rem;
            border: 1px solid var(--border);
            border-radius: 8px;
            background: var(--surface);
            box-shadow: 0 10px 28px var(--shadow);
        }
        .app-brand {
            display: flex;
            align-items: center;
            gap: 0.8rem;
            min-width: 0;
        }
        .app-logo {
            width: 44px;
            height: 44px;
            object-fit: contain;
            border-radius: 8px;
            background: #ffffff;
            border: 1px solid var(--border);
            padding: 0.25rem;
        }
        .app-title {
            color: var(--text);
            font-size: clamp(1.55rem, 3vw, 2.25rem);
            line-height: 1.1;
            font-weight: 850;
            margin: 0;
        }
        .app-subtitle {
            color: var(--text-soft);
            font-size: 0.95rem;
            margin-top: 0.25rem;
        }
        .user-box {
            max-width: 18rem;
            color: var(--text-soft);
            font-size: 0.88rem;
            text-align: right;
            overflow: hidden;
            text-overflow: ellipsis;
            white-space: nowrap;
        }
        .panel-title {
            margin: 0 0 0.25rem;
            color: var(--text);
            font-size: 1.25rem;
            font-weight: 800;
        }
        .panel-copy {
            margin: 0 0 1rem;
            color: var(--text-soft);
            font-size: 0.95rem;
        }
        [data-testid="stVerticalBlockBorderWrapper"] {
            padding: 1rem;
            border: 1px solid var(--border) !important;
            border-radius: 8px !important;
            background: var(--surface) !important;
            box-shadow: 0 8px 22px var(--shadow);
        }
        div.stButton > button,
        div.stDownloadButton > button,
        [data-testid="stFileUploader"] button {
            border-radius: 8px;
            font-weight: 700;
            min-height: 2.5rem;
        }
        div.stButton > button[kind="primary"],
        div.stDownloadButton > button {
            background: var(--brand);
            border: 1px solid var(--brand);
            color: #ffffff;
        }
        div.stButton > button[kind="primary"]:hover,
        div.stDownloadButton > button:hover {
            background: var(--brand-strong);
            border-color: var(--brand-strong);
            color: #ffffff;
        }
        div.stButton > button:not([kind="primary"]),
        [data-testid="stFileUploader"] button {
            background: var(--surface);
            border: 1px solid var(--border-strong);
            color: var(--brand);
        }
        div.stButton > button:disabled,
        div.stButton > button[kind="primary"]:disabled {
            background: var(--surface-muted) !important;
            border: 1px solid var(--border) !important;
            color: var(--text-muted) !important;
            opacity: 1 !important;
        }
        [data-testid="stFileUploader"] {
            padding: 0;
            background: transparent;
            border: 0;
        }
        [data-testid="stFileUploader"] section {
            min-height: 6rem;
            border: 1px dashed var(--border-strong);
            border-radius: 8px;
            background: var(--surface-muted);
        }
        [data-testid="stFileUploader"] section * {
            color: var(--text) !important;
        }
        [data-testid="stFileUploaderFile"],
        [data-testid="stFileUploader"] [class*="uploadedFile"] {
            display: none !important;
        }
        .selected-files {
            display: grid;
            gap: 0.5rem;
            margin: 0.8rem 0 0.3rem;
        }
        .selected-file {
            display: flex;
            align-items: center;
            gap: 0.75rem;
            padding: 0.75rem;
            border: 1px solid var(--border);
            border-radius: 8px;
            background: var(--surface-muted);
        }
        .selected-file-icon {
            display: grid;
            place-items: center;
            width: 2.4rem;
            height: 2.4rem;
            border-radius: 8px;
            background: var(--brand);
            color: #ffffff;
            font-weight: 800;
            flex: 0 0 auto;
        }
        .selected-file-name {
            color: var(--text);
            font-weight: 750;
            overflow-wrap: anywhere;
        }
        .selected-file-size {
            margin-top: 0.15rem;
            color: var(--text-muted);
            font-size: 0.82rem;
        }
        [data-testid="stExpander"],
        [data-testid="stAlert"] {
            background: var(--surface);
            border-color: var(--border) !important;
            border-radius: 8px;
        }
        [data-testid="stAlert"] * {
            color: var(--text) !important;
        }
        div[data-testid="stAlert"][kind="warning"] {
            background: var(--warning-bg);
        }
        div[data-testid="stAlert"][kind="error"] {
            background: var(--error-bg);
        }
        div[data-testid="stAlert"][kind="success"] {
            background: var(--success-bg);
        }
        @media (max-width: 720px) {
            .app-header {
                align-items: flex-start;
                flex-direction: column;
            }
            .user-box {
                max-width: 100%;
                text-align: left;
            }
        }
        </style>
        """.replace("__THEME_VARS__", theme_vars),
        unsafe_allow_html=True,
    )
def _render_login_page() -> None:
    _render_global_styles()
    google_configured = _google_auth_configured()
    with st.container():
        st.markdown('<div class="auth-shell">', unsafe_allow_html=True)
        if LOGO_PATH.exists():
            st.image(str(LOGO_PATH), width=120)
        _render_theme_picker("login_theme")
        st.markdown(
            """
            <div class="auth-kicker">Restricted profile login</div>
            <div class="auth-title">Handwritten Notes Automation</div>
            <div class="auth-copy">
                Sign in with your Google account to generate PW-style handwritten notes.
                Access is limited to approved email addresses from the app allowlist.
            </div>
            """,
            unsafe_allow_html=True,
        )

        if google_configured:
            if st.button("Continue with Google", type="primary", use_container_width=True):
                try:
                    st.login("google")
                except Exception as exc:
                    st.error("Google login failed. Check `.streamlit/secrets.toml` and restart the app.")
                    st.caption(str(exc))
        else:
            st.error(
                "Google OAuth is not configured yet. Replace the placeholder "
                "`client_id` and `client_secret` in `.streamlit/secrets.toml` "
                "with real Google OAuth credentials, then restart the app."
            )

        st.markdown("</div>", unsafe_allow_html=True)


def _require_allowed_google_user() -> None:
    is_logged_in = bool(getattr(st.user, "is_logged_in", False))
    if not is_logged_in:
        _render_login_page()
        st.stop()

    if _auth_session_expired():
        _logout_user()
        st.stop()

    allowed_emails = _load_allowed_emails()
    email = _current_user_email()
    if not email or email not in allowed_emails:
        _render_global_styles()
        st.error("This Google account is not allowed to use this app.")
        if email:
            st.caption(f"Signed in as {email}")
        if st.button("Sign out", use_container_width=False):
            _logout_user()
        st.stop()


if "ui_theme" not in st.session_state:
    st.session_state.ui_theme = "Light"
_require_allowed_google_user()
_render_global_styles()
user_email = _current_user_email()
user_initial = (user_email[:1] or "U").upper()
logo_html = ""
if LOGO_PATH.exists():
    logo_data = base64.b64encode(LOGO_PATH.read_bytes()).decode("ascii")
    logo_html = f'<img class="app-brand-logo" src="data:image/png;base64,{logo_data}" alt="PW logo">'

with st.sidebar:
    if LOGO_PATH.exists():
        st.image(str(LOGO_PATH), width=92)
    _render_theme_picker("sidebar_theme")
    st.markdown("### Profile")
    st.caption(_current_user_email())
    if st.button("Sign out", key="sidebar_sign_out", use_container_width=True):
        _logout_user()

st.markdown(
    f"""
    <div class="app-header">
        <div class="app-brand">
            {logo_html.replace("app-brand-logo", "app-logo")}
            <div>
                <div class="app-title">Handwritten Notes Automation</div>
                <div class="app-subtitle">Upload a PDF or PowerPoint and download generated notes.</div>
            </div>
        </div>
        <div class="user-box">Signed in as<br>{user_email}</div>
    </div>
    """,
    unsafe_allow_html=True,
)
# ---- Fixed configuration (no user-facing settings) -------------------------
LANGUAGE = "English"
MODE = "summary"
SUBJECT = "auto"            # auto-detected from the file
SEND_IMAGES = True
STRICT_FILTER = True
DTP_POLICY = "hide_note_insert_image"
IMAGE_MODE = "smart_crop"
AI_REDRAW = True
IMAGE_MODEL = "gemini-2.5-flash-image"

if "results" not in st.session_state:
    st.session_state.results = []
if "uploader_key" not in st.session_state:
    st.session_state.uploader_key = 0


def _result_to_dict(name: str, result) -> dict:
    notes = ""
    if result.raw_notes_path and result.raw_notes_path.exists():
        notes = result.raw_notes_path.read_text(encoding="utf-8")
    return {
        "name": name,
        "run_dir": str(result.run_dir),
        "docx": str(result.docx_path) if result.docx_path else None,
        "pdf": str(result.pdf_path) if result.pdf_path else None,
        "zip": str(result.zip_path) if result.zip_path else None,
        "notes": notes,
        "tracking": None,
        "error": None,
    }


def _notes_to_markdown(notes: str) -> str:
    out: list[str] = []
    for raw in notes.splitlines():
        line = raw.rstrip()
        if not line.strip():
            out.append("")
            continue
        leading = len(raw) - len(raw.lstrip(" \t"))
        level = min(3, leading // 4)
        content = line.strip()
        bullet_m = re.match(r"^(\*|•|·|-|–)\s+(.*)$", content)
        if bullet_m:
            out.append("  " * level + "- " + bullet_m.group(2).strip())
        elif content[0].isdigit() and content[1:2] in {".", ")"}:
            out.append("  " * level + content)
        elif "note to dtp" in content.lower():
            continue  # hidden in output, so hide in preview too
        else:
            out.append(content)
    return "\n".join(out)


def _format_file_size(size: int) -> str:
    units = ["B", "KB", "MB", "GB"]
    value = float(size)
    for unit in units:
        if value < 1024 or unit == units[-1]:
            return f"{value:.1f} {unit}" if unit != "B" else f"{int(value)} {unit}"
        value /= 1024
    return f"{size} B"


def _render_selected_files(files) -> None:
    if not files:
        return
    rows = []
    for uploaded in files:
        name = html.escape(uploaded.name)
        size = html.escape(_format_file_size(uploaded.size))
        suffix = html.escape(Path(uploaded.name).suffix.replace(".", "").upper() or "FILE")
        rows.append(
            f"""
            <div class="selected-file">
                <div class="selected-file-icon">{suffix[:4]}</div>
                <div>
                    <div class="selected-file-name">{name}</div>
                    <div class="selected-file-size">{size}</div>
                </div>
            </div>
            """
        )
    st.markdown('<div class="selected-files">' + "".join(rows) + "</div>", unsafe_allow_html=True)


with st.container(border=True):
    st.markdown(
        """
        <div class="panel-title">Upload file</div>
        <div class="panel-copy">Choose PDF, PPTX, or PPT files to generate notes.</div>
        """,
        unsafe_allow_html=True,
    )
    uploaded_files = st.file_uploader(
        "Upload lecture file(s)",
        type=["pdf", "pptx", "ppt"],
        accept_multiple_files=True,
        label_visibility="collapsed",
        key=f"lecture_files_{st.session_state.uploader_key}",
    )
    if uploaded_files:
        st.markdown(
            """
            <style>
            div[data-testid="stFileUploader"] {
                display: none !important;
            }
            </style>
            """,
            unsafe_allow_html=True,
        )
        _render_selected_files(uploaded_files)
        if st.button("Change files", key="change_uploaded_files", use_container_width=False):
            st.session_state.uploader_key += 1
            st.rerun()
    run_button = st.button("Generate handwritten notes", type="primary", disabled=not uploaded_files)

if run_button and uploaded_files:
    if not GEMINI_API_KEY:
        st.error("No Gemini API key configured. Add GEMINI_API_KEY to the .env file.")
    else:
        st.session_state.results = []
        total = len(uploaded_files)
        progress = st.progress(0.0, text=f"Starting {total} file(s)...")
        with tempfile.TemporaryDirectory() as tmpdir:
            for i, uploaded in enumerate(uploaded_files, start=1):
                progress.progress((i - 1) / total, text=f"[{i}/{total}] {uploaded.name}")
                tmp_path = Path(tmpdir) / uploaded.name
                tmp_path.write_bytes(uploaded.getbuffer())
                try:
                    result = run_pipeline(
                        tmp_path, subject=SUBJECT, language=LANGUAGE, mode=MODE,
                        api_key=GEMINI_API_KEY, model=DEFAULT_MODEL, provider=DEFAULT_PROVIDER,
                        send_images_to_ai=SEND_IMAGES, strict_filter=STRICT_FILTER,
                        allow_mock=False, image_insert_mode=IMAGE_MODE, dtp_note_policy=DTP_POLICY,
                        ai_redraw_diagrams=AI_REDRAW, image_model=IMAGE_MODEL,
                    )
                    result_dict = _result_to_dict(uploaded.name, result)
                    result_dict["tracking"] = _track_usage_for_result(uploaded.name, result)
                    st.session_state.results.append(result_dict)
                except GeminiError as e:
                    st.session_state.results.append({"name": uploaded.name, "error": f"Gemini API failed: {e}", "notes": "", "run_dir": None, "tracking": None})
                except Exception as e:
                    st.session_state.results.append({"name": uploaded.name, "error": f"Processing failed: {e}", "notes": "", "run_dir": None, "tracking": None})
        progress.progress(1.0, text="Done")
        ok = sum(1 for r in st.session_state.results if not r.get("error"))
        st.success(f"Completed: {ok}/{total} file(s).")


results = st.session_state.results
if results:
    st.divider()
    for idx, r in enumerate(results):
        with st.expander(r["name"] + ("  ❌" if r.get("error") else ""), expanded=(len(results) == 1 or bool(r.get("error")))):
            if r.get("error"):
                st.error(r["error"])
                continue
            if r.get("tracking"):
                tracking = str(r["tracking"])
                if "not updated" in tracking.lower() or "failed" in tracking.lower():
                    st.warning(f"Usage saved locally, but spreadsheet sync failed: {tracking}")
                else:
                    st.caption(f"Usage tracked via {tracking}.")
            c1, c2, c3 = st.columns(3)
            with c1:
                if r["docx"] and Path(r["docx"]).exists():
                    st.download_button("Download DOCX", Path(r["docx"]).read_bytes(), file_name=Path(r["docx"]).name, key=f"docx_{idx}")
            with c2:
                if r["pdf"] and Path(r["pdf"]).exists():
                    st.download_button("Download PDF", Path(r["pdf"]).read_bytes(), file_name=Path(r["pdf"]).name, key=f"pdf_{idx}")
            with c3:
                if r["zip"] and Path(r["zip"]).exists():
                    st.download_button("Download ZIP", Path(r["zip"]).read_bytes(), file_name=Path(r["zip"]).name, key=f"zip_{idx}")

            tab_preview, tab_images = st.tabs(["Preview", "Diagrams"])
            with tab_preview:
                st.markdown(_notes_to_markdown(r["notes"]))
            with tab_images:
                run_dir = Path(r["run_dir"]) if r["run_dir"] else None
                imgs = []
                if run_dir and (run_dir / "ai_diagrams").exists():
                    imgs = sorted((run_dir / "ai_diagrams").glob("*_t.png")) or sorted((run_dir / "ai_diagrams").glob("*.png"))
                if not imgs and run_dir and (run_dir / "inserted_images").exists():
                    imgs = sorted((run_dir / "inserted_images").glob("*.png"))
                if imgs:
                    cols = st.columns(2)
                    for j, img in enumerate(imgs):
                        with cols[j % 2]:
                            st.image(str(img), use_container_width=True)
                else:
                    st.caption("No diagrams inserted for this file.")
